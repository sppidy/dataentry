import tkinter
import openpyxl
import os
from tkinter import ttk
from tkinter import messagebox
# Path: program.py
basepath = os.path.dirname(__file__)
def main():
    #-------------------------------------------------------- Window Configuration ----------------------------------------------------------#
    
    window = tkinter.Tk()
    window.title("Data Entry Form")
    window.configure(background="#CCCCFF")
    window.geometry("1280x720")
    style= ttk.Style()
    style.theme_use("alt")
    style.theme_settings("alt",{"TLabelFrame":{"configure":{"background":"#CCCCFF"}}})
    style.configure("TFrame", background= "#CCCCFF", foreground= "black", font=("Calibiri", 10))
    style.configure("TCombobox", fieldbackground= "light blue", background= "light blue", foreground= "black", bordercolor= "black", arrowcolor= "black", selectbackground= "white", selectforeground= "black", font=("Calibiri"))
    style.configure("TEntry", fieldbackground= "light blue", background= "light blue", foreground= "black", bordercolor= "black", selectbackground= "white", selectforeground= "black", font=("Calibiri"))
    style.configure("TLabel", background= "#CCCCFF", foreground= "black", font=("Calibiri", 11, 'bold', 'italic'))
    style.configure("TLabelFrame",bordercolor= "black", background= "#CCCCFF", foreground= "black", font=("Calibiri", 10))
    style.configure("TLabelFrame.Label", background= "#CCCCFF", foreground= "black", font=("Calibiri", 10))
    style.configure("data_info", background= "#CCCCFF", foreground= "black", font=("Calibiri", 10))
    style.configure("data1_info", background= "#CCCCFF", foreground= "black", font=("Calibiri", 10))
    frame = ttk.Frame(window)
    frame.pack()
    canvas=tkinter.Canvas(frame, width=1280, height=1024, background="#CCCCFF")
    scroll_bar=ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
    scroll_bar.pack(side="right", fill="y")
    scroll_bar1=ttk.Scrollbar(frame, orient="horizontal", command=canvas.xview)
    scroll_bar1.pack(side="bottom", fill="x")
    canvas.pack(ipadx=40,ipady=10, fill="both", expand=1)
    canvas.configure(yscrollcommand=scroll_bar.set, xscrollcommand=scroll_bar1.set)
    canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    second_frame=ttk.Frame(canvas)
    canvas.create_window((0,0), window=second_frame, anchor="nw")
    #-------------------------------------------------------- Diesel and RunTime ----------------------------------------------------------#
    # Saving Info
    data_info =tkinter.LabelFrame(second_frame, text="Diesel and Runtime Entry",background="#CCCCFF", foreground="black", font=("Calibiri", 10))
    data_info.grid(row= 0, column=0, padx=20, pady=10)
    date_label = ttk.Label(data_info, text="Date")
    date_label.grid(row=0, column=0)
    date_entry = ttk.Entry(data_info)
    date_entry.grid(row=1, column=0)
    bunk_name_label = ttk.Label(data_info, text="Bulk Name")
    bunk_name_label.grid(row=0, column=1)
    bunk_name = ttk.Combobox(data_info, values=["Thirupathi","Saravana","Nayara","Others"])
    bunk_name.grid(row=1, column=1)
    diesel_purchased_label = ttk.Label(data_info, text="Diesel Purchased")
    diesel_purchased_label.grid(row=0, column=2)
    diesel_purchased=ttk.Combobox(data_info, values=[100,200,400,600,800,1000,1200,1400])
    diesel_purchased.grid(row=1, column=2)
    sanybreakerhours_label = ttk.Label(data_info, text="Sany Breaker Hours")
    sanybreakerhours_label.grid(row=0, column=3)
    sanybreakerhours_entry = ttk.Entry(data_info)
    sanybreakerhours_entry.grid(row=1, column=3)
    sanybuckerhours_label = ttk.Label(data_info, text="Sany Bucket Hours")
    sanybuckerhours_label.grid(row=0, column=4)
    sanybuckerhours_entry = ttk.Entry(data_info)
    sanybuckerhours_entry.grid(row=1, column=4)
    volvo240buckethours_label = ttk.Label(data_info, text="Volvo 240 Bucket Hours")
    volvo240buckethours_label.grid(row=0, column=5)
    volvo240buckethours_entry = ttk.Entry(data_info)
    volvo240buckethours_entry.grid(row=1, column=5)
    volvo210buckethours_label = ttk.Label(data_info, text="Volvo 210 Breaker Hours")
    volvo210buckethours_label.grid(row=2, column=0)
    volvo210buckethours_entry = ttk.Entry(data_info)
    volvo210buckethours_entry.grid(row=3, column=0)
    car_diesel_ltrs_label = ttk.Label(data_info, text="Car Diesel Ltrs")
    car_diesel_ltrs_label.grid(row=2, column=1)
    car_diesel_ltrs_entry = ttk.Entry(data_info)
    car_diesel_ltrs_entry.grid(row=3, column=1)
    _3359diesel_label = ttk.Label(data_info, text="3359 Diesel Ltrs")
    _3359diesel_label.grid(row=2, column=2)
    _3359diesel_entry = ttk.Entry(data_info)
    _3359diesel_entry.grid(row=3, column=2)
    _770_diesel_label = ttk.Label(data_info, text="770 Diesel Ltrs")
    _770_diesel_label.grid(row=2, column=3)
    _770_diesel_entry = ttk.Entry(data_info)
    _770_diesel_entry.grid(row=3, column=3)
    _compressor1diesel_label = ttk.Label(data_info, text="Compressor1 Diesel Ltrs")
    _compressor1diesel_label.grid(row=2, column=4)
    _compressor1diesel_entry = ttk.Entry(data_info)
    _compressor1diesel_entry.grid(row=3, column=4)
    _compressor2diesel_label = ttk.Label(data_info, text="Compressor2 Diesel Ltrs")
    _compressor2diesel_label.grid(row=2, column=5)
    _compressor2diesel_entry = ttk.Entry(data_info)
    _compressor2diesel_entry.grid(row=3, column=5)
    _volvo240diesel_label = ttk.Label(data_info, text="Volvo 240 Diesel Ltrs")
    _volvo240diesel_label.grid(row=4, column=0)
    _volvo240diesel_entry = ttk.Entry(data_info)
    _volvo240diesel_entry.grid(row=5, column=0)
    _volvo210diesel_label = ttk.Label(data_info, text="Volvo 210 Diesel Ltrs")
    _volvo210diesel_label.grid(row=4, column=1)
    _volvo210diesel_entry = ttk.Entry(data_info)
    _volvo210diesel_entry.grid(row=5, column=1)
    _sany140diesel_label = ttk.Label(data_info, text="Sany 140 Diesel Ltrs")
    _sany140diesel_label.grid(row=4, column=2)
    _sany140diesel_entry = ttk.Entry(data_info)
    _sany140diesel_entry.grid(row=5, column=2)
    _diesel_price_on_day_label = ttk.Label(data_info, text="Diesel Price On Day")
    _diesel_price_on_day_label.grid(row=4, column=3)
    _diesel_price_on_day_entry = ttk.Entry(data_info)
    _diesel_price_on_day_entry.grid(row=5, column=3)
    _5569_loads_label = ttk.Label(data_info, text="5569 Loads")
    _5569_loads_label.grid(row=4, column=4)
    _5569_loads_entry = ttk.Entry(data_info)
    _5569_loads_entry.grid(row=5, column=4)
    _9970_loads_label = ttk.Label(data_info, text="9970 Loads")
    _9970_loads_label.grid(row=4, column=5)
    _9970_loads_entry = ttk.Entry(data_info)
    _9970_loads_entry.grid(row=5, column=5)
    _222_loads_label = ttk.Label(data_info, text="222 Loads")
    _222_loads_label.grid(row=6, column=0)
    _222_loads_entry = ttk.Entry(data_info)
    _222_loads_entry.grid(row=7, column=0)
    _3359_loads_label = ttk.Label(data_info, text="3359 Loads")
    _3359_loads_label.grid(row=6, column=1)
    _3359_loads_entry = ttk.Entry(data_info)
    _3359_loads_entry.grid(row=7, column=1)
    _5506_loads_label = ttk.Label(data_info, text="5506 Loads")
    _5506_loads_label.grid(row=6, column=2)
    _5506_loads_entry = ttk.Entry(data_info)
    _5506_loads_entry.grid(row=7, column=2)
    _5508_loads_label = ttk.Label(data_info, text="5508 Loads")
    _5508_loads_label.grid(row=6, column=3)
    _5508_loads_entry = ttk.Entry(data_info)
    _5508_loads_entry.grid(row=7, column=3)
    _950_loads_label = ttk.Label(data_info, text="950 Loads")
    _950_loads_label.grid(row=6, column=4)
    _950_loads_entry = ttk.Entry(data_info)
    _950_loads_entry.grid(row=7, column=4)

    for widget in data_info.winfo_children():
        widget.grid_configure(padx=7, pady=10)

    def clear():
        date_entry.delete(0, 'end')
        bunk_name.delete(0, 'end')
        diesel_purchased.delete(0, 'end')
        sanybreakerhours_entry.delete(0, 'end')
        sanybuckerhours_entry.delete(0, 'end')
        volvo240buckethours_entry.delete(0, 'end')
        volvo210buckethours_entry.delete(0, 'end')
        car_diesel_ltrs_entry.delete(0, 'end')
        _3359diesel_entry.delete(0, 'end')
        _770_diesel_entry.delete(0, 'end')
        _compressor1diesel_entry.delete(0, 'end')
        _compressor2diesel_entry.delete(0, 'end')
        _volvo240diesel_entry.delete(0, 'end')
        _volvo210diesel_entry.delete(0, 'end')
        _sany140diesel_entry.delete(0, 'end')
        _diesel_price_on_day_entry.delete(0, 'end')
        _5569_loads_entry.delete(0, 'end')
        _9970_loads_entry.delete(0, 'end')
        _222_loads_entry.delete(0, 'end')
        _3359_loads_entry.delete(0, 'end')
        _5506_loads_entry.delete(0, 'end')
        _5508_loads_entry.delete(0, 'end')
        _950_loads_entry.delete(0, 'end')

    filepath = "C:\\Users\\Thirupathi\\OneDrive - MSFT\\Daily Reports\\mpdrentry.xlsx"
    if os.path.exists("C:\\Users\\Thirupathi\\OneDrive - MSFT\\Daily Reports\\")==False:
        os.mkdir("C):\\Users\\Thirupathi\\OneDrive - MSFT\\Daily Reports\\")

    # filepath = "D:\\Excel\\mpdrentry.xlsx"
    # if os.path.exists("D:\\Excel\\")==False:
    #     os.mkdir("D:\\Excel\\")

    def enter_data():
        global filepath
        date=date_entry.get()
        bunkname=bunk_name.get()
        dieselpurchased=int(diesel_purchased.get())
        sanybreakerhours=int(sanybreakerhours_entry.get())
        sanybuckerhours=int(sanybuckerhours_entry.get())
        volvo240breakerhours=int(volvo240buckethours_entry.get())
        volvo210breakerhours=int(volvo210buckethours_entry.get())
        car_diesel_ltrs=int(car_diesel_ltrs_entry.get())
        _3359_diesel_ltrs=int(_3359diesel_entry.get())
        _770_diesel_ltrs=int(_770_diesel_entry.get())
        _compressor1diesel_=int(_compressor1diesel_entry.get())
        _compressor2diesel_=int(_compressor2diesel_entry.get())
        _volvo_240_diesel_ = int(_volvo240diesel_entry.get())
        _volvo_210_diesel_ = int(_volvo210diesel_entry.get())
        _sany_140_diesel_ = int(_sany140diesel_entry.get())
        _diesel_price_on_day_ = int(_diesel_price_on_day_entry.get())
        _5569_loads_ = int(_5569_loads_entry.get())
        _9970_loads_ = int(_9970_loads_entry.get())
        _222_loads_ = int(_222_loads_entry.get())
        _3359_loads_ = int(_3359_loads_entry.get())
        _5506_loads_ = int(_5506_loads_entry.get())
        _5508_loads_ = int(_5508_loads_entry.get())
        _950_loads_ = int(_950_loads_entry.get())
        heading = ["Date", "Bunk Name", "Diesel Purchased", "Sany Breaker Hours", "Sany Bucker Hours", "Volvo 240 Bucket Hours", "Volvo 210 Bucket Hours", "Car Diesel Ltrs", "3359 Diesel Ltrs", "770 Diesel Ltrs", "Compressor 1 Diesel Ltrs", "Compressor 2 Diesel Ltrs", "Volvo 240 Diesel Ltrs", "Volvo 210 Diesel Ltrs", "Sany 140 Diesel Ltrs", "Diesel Price on Day", "5569 Loads", "9970 Loads", "222 Loads", "3359 Loads", "5506 Loads", "5508 Loads", "950 Loads"]
        datas=[date, bunkname, dieselpurchased, sanybreakerhours, sanybuckerhours, volvo240breakerhours, volvo210breakerhours, car_diesel_ltrs, _3359_diesel_ltrs, _770_diesel_ltrs, _compressor1diesel_, _compressor2diesel_, _volvo_240_diesel_, _volvo_210_diesel_, _sany_140_diesel_, _diesel_price_on_day_, _5569_loads_, _9970_loads_, _222_loads_, _3359_loads_, _5506_loads_, _5508_loads_, _950_loads_]
        if any(field == '' or field == ' ' for field in datas):
            messagebox.showerror("Error", "Please Fill All the Fields.")
        else:
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                workbook.create_sheet("Diesel and Runtime")
                sheet=workbook.active=workbook["Diesel and Runtime"]
                sheet.append(heading)
                sheet.append(datas)
                workbook.save(filepath)
                clear()
                messagebox.showinfo("Success","FCreated and Data Submitted Successfully")
            else:
                workbook = openpyxl.load_workbook(filepath)
                #CHECK IF SHEET EXISTS OR NOT
                if "Diesel and Runtime" not in workbook.sheetnames:
                    workbook.create_sheet("Diesel and Runtime")
                    sheet=workbook["Diesel and Runtime"]
                    sheet.append(heading)
                    sheet.append(datas)
                    workbook.save(filepath)
                    clear()
                    messagebox.showinfo("Success","SCreated and Submitted Succesfully.")
                else:
                    sheet=workbook["Diesel and Runtime"]
                    sheet.append(datas)
                    workbook.save(filepath)
                    clear()
                    messagebox.showinfo("Success","Data Submitted Successfully")

    def delete():
        global filepath
        window2 = tkinter.Tk()
        window2.title("Delete Data")
        frame1 = ttk.Frame(window2,padding=20)
        frame1.pack()
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook["Diesel and Runtime"]
        def delete_data():
            date_delete = date_delete_entry.get()
            found = False
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == date_delete:
                    sheet.delete_rows(i)
                    found = True
                    break
            workbook.save(filepath)
            window2.destroy()
            if found:
                messagebox.showinfo("Success", "Data Deleted Successfully")
            else:
                messagebox.showerror("Error", "Data Not Found")
        dlabel = ttk.Label(frame1, text="Enter Date to Delete", background="#CCCCFF")
        dlabel.grid(row=0, column=0)
        date_delete_entry = ttk.Entry(frame1)
        date_delete_entry.grid(row=0, column=1)
        button_delete1 = ttk.Button(frame1, text="Delete", command=delete_data)
        button_delete1.grid(row=1, column=0)
        window2.iconbitmap(basepath+"\\906310.ico")
        window2.mainloop()


    def edit():
        window1 = tkinter.Tk()
        window1.title("Edit Data")
        window1.configure(background="#CCCCFF")
        frame1 = ttk.Frame(window1,padding=20)
        frame1.pack()
        dlabel = ttk.Label(frame1, text="Enter Date to Edit", background="#CCCCFF")
        dlabel.grid(row=0, column=0)
        date_edit_entry = ttk.Entry(frame1)
        date_edit_entry.grid(row=0, column=1)
        def edit_data():
            date_edit = date_edit_entry.get()
            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook["Diesel and Runtime"]
            found = False
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == date_edit:
                    found = True
                    date_entry.insert(0, sheet.cell(row=i, column=1).value)
                    bunk_name.insert(0, sheet.cell(row=i, column=2).value)
                    diesel_purchased.insert(0, sheet.cell(row=i, column=3).value)
                    sanybreakerhours_entry.insert(0, sheet.cell(row=i, column=4).value)
                    sanybuckerhours_entry.insert(0, sheet.cell(row=i, column=5).value)
                    volvo240buckethours_entry.insert(0, sheet.cell(row=i, column=6).value)
                    volvo210buckethours_entry.insert(0, sheet.cell(row=i, column=7).value)
                    car_diesel_ltrs_entry.insert(0, sheet.cell(row=i, column=8).value)
                    _3359diesel_entry.insert(0, sheet.cell(row=i, column=9).value)
                    _770_diesel_entry.insert(0, sheet.cell(row=i, column=10).value)
                    _compressor1diesel_entry.insert(0, sheet.cell(row=i, column=11).value)
                    _compressor2diesel_entry.insert(0, sheet.cell(row=i, column=12).value)
                    _volvo240diesel_entry.insert(0, sheet.cell(row=i, column=13).value)
                    _volvo210diesel_entry.insert(0, sheet.cell(row=i, column=14).value)
                    _sany140diesel_entry.insert(0, sheet.cell(row=i, column=15).value)
                    _diesel_price_on_day_entry.insert(0, sheet.cell(row=i, column=16).value)
                    _5569_loads_entry.insert(0, sheet.cell(row=i, column=17).value)
                    _9970_loads_entry.insert(0, sheet.cell(row=i, column=18).value)
                    _222_loads_entry.insert(0, sheet.cell(row=i, column=19).value)
                    _3359_loads_entry.insert(0, sheet.cell(row=i, column=20).value)
                    _5506_loads_entry.insert(0, sheet.cell(row=i, column=21).value)
                    _5508_loads_entry.insert(0, sheet.cell(row=i, column=22).value)
                    _950_loads_entry.insert(0, sheet.cell(row=i, column=23).value)
                    break
            workbook.save(filepath)
            window1.destroy()
            if found:
                messagebox.showinfo("Success", "Data Found and Filled in the Form.")
            else:
                messagebox.showinfo("Error", "Data Not Found")
        button_edit1 = ttk.Button(frame1, text="Edit", command=edit_data)
        button_edit1.grid(row=1, column=0)
        window1.iconbitmap(basepath+"\\906310.ico")
        window1.mainloop()



    def modify():
        global window1
        global filepath
        workbook = openpyxl.load_workbook(filepath)
        sheet=workbook.active=workbook["Diesel and Runtime"]
        date=date_entry.get()
        bunkname=bunk_name.get()
        dieselpurchased=float(diesel_purchased.get())
        sanybreakerhours=float(sanybreakerhours_entry.get())
        sanybuckerhours=float(sanybuckerhours_entry.get())
        volvo240breakerhours=float(volvo240buckethours_entry.get())
        volvo210breakerhours=float(volvo210buckethours_entry.get())
        car_diesel_ltrs=float(car_diesel_ltrs_entry.get())
        _3359_diesel_ltrs=float(_3359diesel_entry.get())
        _770_diesel_ltrs=float(_770_diesel_entry.get())
        _compressor1diesel_=float(_compressor1diesel_entry.get())
        _compressor2diesel_=float(_compressor2diesel_entry.get())
        _volvo_240_diesel_ = float(_volvo240diesel_entry.get())
        _volvo_210_diesel_ = float(_volvo210diesel_entry.get())
        _sany_140_diesel_ = float(_sany140diesel_entry.get())
        _diesel_price_on_day_ = float(_diesel_price_on_day_entry.get())
        _5569_loads_ = float(_5569_loads_entry.get())
        _9970_loads_ = float(_9970_loads_entry.get())
        _222_loads_ = float(_222_loads_entry.get())
        _3359_loads_ = float(_3359_loads_entry.get())
        _5506_loads_ = float(_5506_loads_entry.get())
        _5508_loads_ = float(_5508_loads_entry.get())
        _950_loads_ = float(_950_loads_entry.get())
        if any(field == '' or field == ' ' for field in [date, bunkname, dieselpurchased, sanybreakerhours, sanybuckerhours, volvo240breakerhours, volvo210breakerhours, car_diesel_ltrs, _3359_diesel_ltrs, _770_diesel_ltrs, _compressor1diesel_, _compressor2diesel_, _volvo_240_diesel_, _volvo_210_diesel_, _sany_140_diesel_, _diesel_price_on_day_, _5569_loads_, _9970_loads_, _222_loads_, _3359_loads_, _5506_loads_, _5508_loads_, _950_loads_]):
            messagebox.showerror("Error", "Please Fill All the Fields.")
        else:
            for i in range(1,sheet.max_row+1):
                if sheet.cell(row=i,column=1).value==date:
                    sheet.cell(row=i,column=2).value=bunkname
                    sheet.cell(row=i,column=3).value=dieselpurchased
                    sheet.cell(row=i,column=4).value=sanybreakerhours
                    sheet.cell(row=i,column=5).value=sanybuckerhours
                    sheet.cell(row=i,column=6).value=volvo240breakerhours
                    sheet.cell(row=i,column=7).value=volvo210breakerhours
                    sheet.cell(row=i,column=8).value=car_diesel_ltrs
                    sheet.cell(row=i,column=9).value=_3359_diesel_ltrs
                    sheet.cell(row=i,column=10).value=_770_diesel_ltrs
                    sheet.cell(row=i,column=11).value=_compressor1diesel_
                    sheet.cell(row=i,column=12).value=_compressor2diesel_
                    sheet.cell(row=i,column=13).value=_volvo_240_diesel_
                    sheet.cell(row=i,column=14).value=_volvo_210_diesel_
                    sheet.cell(row=i,column=15).value=_sany_140_diesel_
                    sheet.cell(row=i,column=16).value=_diesel_price_on_day_
                    sheet.cell(row=i,column=17).value=_5569_loads_
                    sheet.cell(row=i,column=18).value=_9970_loads_
                    sheet.cell(row=i,column=19).value=_222_loads_
                    sheet.cell(row=i,column=20).value=_3359_loads_
                    sheet.cell(row=i,column=21).value=_5506_loads_
                    sheet.cell(row=i,column=22).value=_5508_loads_
                    sheet.cell(row=i,column=23).value=_950_loads_
                    workbook.save(filepath)
                    messagebox.showinfo("Success","Data Modified Successfully.")
                    clear()
                    break

    button_frame = ttk.Frame(second_frame)
    button_frame.grid(row= 1, column=0, padx=40, pady=10)
        
    # Button Save
    button_save = tkinter.Button(button_frame, text="ㅤㅤㅤSaveㅤㅤㅤ", command= enter_data)
    button_save.configure(background="green")
    button_save.grid(row=0, column=0, sticky="news")
    # Button Clear
    button_clear = tkinter.Button(button_frame, text="ㅤㅤㅤClearㅤㅤㅤ", command= clear)
    button_clear.configure(background="yellow")
    button_clear.grid(row=0, column=1, sticky="news")
    # Button Edit
    button_edit = tkinter.Button(button_frame, text="ㅤㅤㅤEditㅤㅤㅤ", command= edit)
    button_edit.configure(background="cyan")
    button_edit.grid(row=0, column=2, sticky="news")
    #Button Delete
    button_delete = tkinter.Button(button_frame, text="ㅤㅤㅤDeleteㅤㅤㅤ", command= delete)
    button_delete.configure(background="red")
    button_delete.grid(row=0, column=4, sticky="news")
    #Button Modify
    button_modify=tkinter.Button(button_frame,text="ㅤㅤㅤModifyㅤㅤㅤ",command=modify)
    button_modify.configure(background="orange")
    button_modify.grid(row=0,column=3,sticky="news")
    #-----------------------------------------------------End of Diesel and RunTime -------------------------------------------------------#
    #----------------------------------------------------------Explosives------------------------------------------------------------------#
    data1_info =tkinter.LabelFrame(second_frame, text="Mines Entry",background="#CCCCFF",font=("Calibiri", 10))
    data1_info.grid(row= 2, column=0, padx=40, pady=10)
    date1_label=ttk.Label(data1_info, text="Date",background="#CCCCFF")
    date1_label.grid(row=0, column=0)
    date1_entry=ttk.Entry(data1_info)
    date1_entry.grid(row=1, column=0)
    blasting_done_on_label_=ttk.Label(data1_info, text="Blasting Done On")
    blasting_done_on_label_.grid(row=0, column=1)
    blasting_done_on_entry_=ttk.Combobox(data1_info, values=["Soft Rock", "Hard Rock"])
    blasting_done_on_entry_.grid(row=1, column=1)
    _3m_excel_label_=ttk.Label(data1_info, text="3m Excel Reciept")
    _3m_excel_label_.grid(row=0, column=2)
    _3m_excel_entry_=ttk.Entry(data1_info)
    _3m_excel_entry_.grid(row=1, column=2)
    _3m_excel_used_label_=ttk.Label(data1_info, text="3m Excel Used")
    _3m_excel_used_label_.grid(row=0, column=3)
    _3m_excel_used_entry_=ttk.Entry(data1_info)
    _3m_excel_used_entry_.grid(row=1, column=3)
    _4m_reciepts_label_=ttk.Label(data1_info, text="4m Excel Reciepts")
    _4m_reciepts_label_.grid(row=0, column=4)
    _4m_reciepts_entry_=ttk.Entry(data1_info)
    _4m_reciepts_entry_.grid(row=1, column=4)
    _4m_excel_used_label_=ttk.Label(data1_info, text="4m Excel Used")
    _4m_excel_used_label_.grid(row=0, column=5)
    _4m_excel_used_entry_=ttk.Entry(data1_info)
    _4m_excel_used_entry_.grid(row=1, column=5)
    _5m_excel_reciepts_label_=ttk.Label(data1_info, text="5m Excel Reciepts")
    _5m_excel_reciepts_label_.grid(row=2, column=0)
    _5m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _5m_excel_reciepts_entry_.grid(row=3, column=0)
    _5m_excel_used_label_=ttk.Label(data1_info, text="5m Excel Used")
    _5m_excel_used_label_.grid(row=2, column=1)
    _5m_excel_used_entry_=ttk.Entry(data1_info)
    _5m_excel_used_entry_.grid(row=3, column=1)
    _6m_excel_reciepts_label_=ttk.Label(data1_info, text="6m Excel Reciepts")
    _6m_excel_reciepts_label_.grid(row=2, column=2)
    _6m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _6m_excel_reciepts_entry_.grid(row=3, column=2)
    _6m_excel_used_label_=ttk.Label(data1_info, text="6m Excel Used")
    _6m_excel_used_label_.grid(row=2, column=3)
    _6m_excel_used_entry_=ttk.Entry(data1_info)
    _6m_excel_used_entry_.grid(row=3, column=3)
    _7m_excel_reciepts_label_=ttk.Label(data1_info, text="7m Excel Reciepts")
    _7m_excel_reciepts_label_.grid(row=2, column=4)
    _7m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _7m_excel_reciepts_entry_.grid(row=3, column=4)
    _7m_excel_used_label_=ttk.Label(data1_info, text="7m Excel Used")
    _7m_excel_used_label_.grid(row=2, column=5)
    _7m_excel_used_entry_=ttk.Entry(data1_info)
    _7m_excel_used_entry_.grid(row=3, column=5)
    _8m_excel_reciepts_label_=ttk.Label(data1_info, text="8m Excel Reciepts")
    _8m_excel_reciepts_label_.grid(row=4, column=0)
    _8m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _8m_excel_reciepts_entry_.grid(row=5, column=0)
    _8m_excel_used_label_=ttk.Label(data1_info, text="8m Excel Used")
    _8m_excel_used_label_.grid(row=4, column=1)
    _8m_excel_used_entry_=ttk.Entry(data1_info)
    _8m_excel_used_entry_.grid(row=5, column=1)
    _9m_excel_reciepts_label_=ttk.Label(data1_info, text="9m Excel Reciepts")
    _9m_excel_reciepts_label_.grid(row=4, column=2)
    _9m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _9m_excel_reciepts_entry_.grid(row=5, column=2)
    _9m_excel_used_label_=ttk.Label(data1_info, text="9m Excel Used")
    _9m_excel_used_label_.grid(row=4, column=3)
    _9m_excel_used_entry_=ttk.Entry(data1_info)
    _9m_excel_used_entry_.grid(row=5, column=3)
    _10m_excel_reciepts_label_=ttk.Label(data1_info, text="10m Excel Reciepts")
    _10m_excel_reciepts_label_.grid(row=4, column=4)
    _10m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _10m_excel_reciepts_entry_.grid(row=5, column=4)
    _10m_excel_used_label_=ttk.Label(data1_info, text="10m Excel Used")
    _10m_excel_used_label_.grid(row=4, column=5)
    _10m_excel_used_entry_=ttk.Entry(data1_info)
    _10m_excel_used_entry_.grid(row=5, column=5)
    _12m_excel_reciepts_label_=ttk.Label(data1_info, text="12m Excel Reciepts")
    _12m_excel_reciepts_label_.grid(row=6, column=0)
    _12m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _12m_excel_reciepts_entry_.grid(row=7, column=0)
    _12m_excel_used_label_=ttk.Label(data1_info, text="12m Excel Used")
    _12m_excel_used_label_.grid(row=6, column=1)
    _12m_excel_used_entry_=ttk.Entry(data1_info)
    _12m_excel_used_entry_.grid(row=7, column=1)
    _15m_excel_reciepts_label_=ttk.Label(data1_info, text="15m Excel Reciepts")
    _15m_excel_reciepts_label_.grid(row=6, column=2)
    _15m_excel_reciepts_entry_=ttk.Entry(data1_info)
    _15m_excel_reciepts_entry_.grid(row=7, column=2)
    _15m_excel_used_label_=ttk.Label(data1_info, text="15m Excel Used")
    _15m_excel_used_label_.grid(row=6, column=3)
    _15m_excel_used_entry_=ttk.Entry(data1_info)
    _15m_excel_used_entry_.grid(row=7, column=3)
    _slurry_reciept_label_=ttk.Label(data1_info, text="Slurry Reciept")
    _slurry_reciept_label_.grid(row=6, column=4)
    _slurry_reciept_entry_=ttk.Entry(data1_info)
    _slurry_reciept_entry_.grid(row=7, column=4)
    _slurry_used_label_=ttk.Label(data1_info, text="Slurry Used")
    _slurry_used_label_.grid(row=6, column=5)
    _slurry_used_entry_=ttk.Entry(data1_info)
    _slurry_used_entry_.grid(row=7, column=5)
    _slurry_big_reciept_label_=ttk.Label(data1_info, text="Slurry Big Reciept")
    _slurry_big_reciept_label_.grid(row=8, column=0)
    _slurry_big_reciept_entry_=ttk.Entry(data1_info)
    _slurry_big_reciept_entry_.grid(row=9, column=0)
    _slurry_big_used_label_=ttk.Label(data1_info, text="Slurry Big Used")
    _slurry_big_used_label_.grid(row=8, column=1)
    _slurry_big_used_entry_=ttk.Entry(data1_info)
    _slurry_big_used_entry_.grid(row=9, column=1)
    _ed_reciept_label_=ttk.Label(data1_info, text="E.D Reciept")
    _ed_reciept_label_.grid(row=8, column=2)
    _ed_reciept_entry_=ttk.Entry(data1_info)
    _ed_reciept_entry_.grid(row=9, column=2)
    _ed_used_label_=ttk.Label(data1_info, text="E.D Used")
    _ed_used_label_.grid(row=8, column=3)
    _ed_used_entry_=ttk.Entry(data1_info)
    _ed_used_entry_.grid(row=9, column=3)
    _d_det_reciepts_label_=ttk.Label(data1_info, text="D.Det Reciepts")
    _d_det_reciepts_label_.grid(row=8, column=4)
    _d_det_reciepts_entry_=ttk.Entry(data1_info)
    _d_det_reciepts_entry_.grid(row=9, column=4)
    _d_det_used_label_=ttk.Label(data1_info, text="D.Det Used")
    _d_det_used_label_.grid(row=8, column=5)
    _d_det_used_entry_=ttk.Entry(data1_info)
    _d_det_used_entry_.grid(row=9, column=5)
    df_5gms_reciepts_label_=ttk.Label(data1_info, text="DF 5gms Reciepts")
    df_5gms_reciepts_label_.grid(row=10, column=0)
    df_5gms_reciepts_entry_=ttk.Entry(data1_info)
    df_5gms_reciepts_entry_.grid(row=11, column=0)
    df_5gms_used_label_=ttk.Label(data1_info, text="DF 5gms Used")
    df_5gms_used_label_.grid(row=10, column=1)
    df_5gms_used_entry_=ttk.Entry(data1_info)
    df_5gms_used_entry_.grid(row=11, column=1)
    _df_10gms_reciepts_label_=ttk.Label(data1_info, text="DF 10gms Reciepts")
    _df_10gms_reciepts_label_.grid(row=10, column=2)
    _df_10gms_reciepts_entry_=ttk.Entry(data1_info)
    _df_10gms_reciepts_entry_.grid(row=11, column=2)
    _df_10gms_used_label_=ttk.Label(data1_info, text="DF 10gms Used")
    _df_10gms_used_label_.grid(row=10, column=3)
    _df_10gms_used_entry_=ttk.Entry(data1_info)
    _df_10gms_used_entry_.grid(row=11, column=3)
    _no_of_holes_2_5_label_=ttk.Label(data1_info, text="No of Holes 2.5")
    _no_of_holes_2_5_label_.grid(row=10, column=4)
    _no_of_holes_2_5_entry_=ttk.Entry(data1_info)
    _no_of_holes_2_5_entry_.grid(row=11, column=4)
    _no_of_holes_5_label_=ttk.Label(data1_info, text="No of Holes 5'")
    _no_of_holes_5_label_.grid(row=10, column=5)
    _no_of_holes_5_entry_=ttk.Entry(data1_info)
    _no_of_holes_5_entry_.grid(row=11, column=5)
    _no_of_holes_6_label_=ttk.Label(data1_info, text="No of Holes 6'")
    _no_of_holes_6_label_.grid(row=12, column=0)
    _no_of_holes_6_entry_=ttk.Entry(data1_info)
    _no_of_holes_6_entry_.grid(row=13, column=0)
    _no_of_holes_8_label_=ttk.Label(data1_info, text="No of Holes 8'")
    _no_of_holes_8_label_.grid(row=12, column=1)
    _no_of_holes_8_entry_=ttk.Entry(data1_info)
    _no_of_holes_8_entry_.grid(row=13, column=1)
    _wagon_drill_4_5_label_=ttk.Label(data1_info, text="Wagon Drill 4.5''")
    _wagon_drill_4_5_label_.grid(row=12, column=2)
    _wagon_drill_4_5_entry_=ttk.Entry(data1_info)
    _wagon_drill_4_5_entry_.grid(row=13, column=2)
    amount_paid_label_=ttk.Label(data1_info, text="Amount Paid to Balan")
    amount_paid_label_.grid(row=12, column=3)
    amount_paid_entry_=ttk.Entry(data1_info)
    amount_paid_entry_.grid(row=13, column=3)


    for widget in data1_info.winfo_children():
        widget.grid_configure(padx=20, pady=10)
        

    def clear1():
        date1_entry.delete(0, 'end')
        blasting_done_on_entry_.delete(0, 'end')
        _3m_excel_entry_.delete(0, 'end')
        _3m_excel_used_entry_.delete(0, 'end')
        _4m_reciepts_entry_.delete(0, 'end')
        _4m_excel_used_entry_.delete(0, 'end')
        _5m_excel_reciepts_entry_.delete(0, 'end')
        _5m_excel_used_entry_.delete(0, 'end')
        _6m_excel_reciepts_entry_.delete(0, 'end')
        _6m_excel_used_entry_.delete(0, 'end')
        _7m_excel_reciepts_entry_.delete(0, 'end')
        _7m_excel_used_entry_.delete(0, 'end')
        _8m_excel_reciepts_entry_.delete(0, 'end')
        _8m_excel_used_entry_.delete(0, 'end')
        _9m_excel_reciepts_entry_.delete(0, 'end')
        _9m_excel_used_entry_.delete(0, 'end')
        _10m_excel_reciepts_entry_.delete(0, 'end')
        _10m_excel_used_entry_.delete(0, 'end')
        _12m_excel_reciepts_entry_.delete(0, 'end')
        _12m_excel_used_entry_.delete(0, 'end')
        _15m_excel_reciepts_entry_.delete(0, 'end')
        _15m_excel_used_entry_.delete(0, 'end')
        _slurry_reciept_entry_.delete(0, 'end')
        _slurry_used_entry_.delete(0, 'end')
        _slurry_big_reciept_entry_.delete(0, 'end')
        _slurry_big_used_entry_.delete(0, 'end')
        _ed_reciept_entry_.delete(0, 'end')
        _ed_used_entry_.delete(0, 'end')
        _d_det_reciepts_entry_.delete(0, 'end')
        _d_det_used_entry_.delete(0, 'end')
        df_5gms_reciepts_entry_.delete(0, 'end')
        df_5gms_used_entry_.delete(0, 'end')
        _df_10gms_reciepts_entry_.delete(0, 'end')
        _df_10gms_used_entry_.delete(0, 'end')
        _no_of_holes_2_5_entry_.delete(0, 'end')
        _no_of_holes_5_entry_.delete(0, 'end')
        _no_of_holes_6_entry_.delete(0, 'end')
        _no_of_holes_8_entry_.delete(0, 'end')
        _wagon_drill_4_5_entry_.delete(0, 'end')
        amount_paid_entry_.delete(0, 'end')


    def save():
        date1=date1_entry.get()
        blasting_done_on=blasting_done_on_entry_.get()
        _3m_excel_=float(_3m_excel_entry_.get())
        _3m_excel_used_=float(_3m_excel_used_entry_.get())
        _4m_reciepts_=float(_4m_reciepts_entry_.get())
        _4m_excel_used_=float(_4m_excel_used_entry_.get())
        _5m_excel_reciepts_=float(_5m_excel_reciepts_entry_.get())
        _5m_excel_used_=float(_5m_excel_used_entry_.get())
        _6m_excel_reciepts_=float(_6m_excel_reciepts_entry_.get())
        _6m_excel_used_=float(_6m_excel_used_entry_.get())
        _7m_excel_reciepts_=float(_7m_excel_reciepts_entry_.get())
        _7m_excel_used_=float(_7m_excel_used_entry_.get())
        _8m_excel_reciepts_=float(_8m_excel_reciepts_entry_.get())
        _8m_excel_used_=float(_8m_excel_used_entry_.get())
        _9m_excel_reciepts_=float(_9m_excel_reciepts_entry_.get())
        _9m_excel_used_=float(_9m_excel_used_entry_.get())
        _10m_excel_reciepts_=float(_10m_excel_reciepts_entry_.get())
        _10m_excel_used_=float(_10m_excel_used_entry_.get())
        _12m_excel_reciepts_=float(_12m_excel_reciepts_entry_.get())
        _12m_excel_used_=float(_12m_excel_used_entry_.get())
        _15m_excel_reciepts_=float(_15m_excel_reciepts_entry_.get())
        _15m_excel_used_=float(_15m_excel_used_entry_.get())
        _slurry_reciept_=float(_slurry_reciept_entry_.get())
        _slurry_used_=float(_slurry_used_entry_.get())
        _slurry_big_reciept_=float(_slurry_big_reciept_entry_.get())
        _slurry_big_used_=float(_slurry_big_used_entry_.get())
        _ed_reciept_=float(_ed_reciept_entry_.get())
        _ed_used_=float(_ed_used_entry_.get())
        _d_det_reciepts_=float(_d_det_reciepts_entry_.get())
        _d_det_used_=float(_d_det_used_entry_.get())
        df_5gms_reciepts_=float(df_5gms_reciepts_entry_.get())
        df_5gms_used_=float(df_5gms_used_entry_.get())
        _df_10gms_reciepts_=float(_df_10gms_reciepts_entry_.get())
        _df_10gms_used_=float(_df_10gms_used_entry_.get())
        _no_of_holes_2_5_=float(_no_of_holes_2_5_entry_.get())
        _no_of_holes_5_=float(_no_of_holes_5_entry_.get())
        _no_of_holes_6_=float(_no_of_holes_6_entry_.get())
        _no_of_holes_8_=float(_no_of_holes_8_entry_.get())
        _wagon_drill_4_5_=float(_wagon_drill_4_5_entry_.get())
        amount_paid_=float(amount_paid_entry_.get())
        heading=["Date", "Blasting Done On", "3M Excel Reciepts", "3M Excel Used", "4M Reciepts", "4M Excel Used","5M Excel Reciepts", "5M Excel Used", "6M Excel Reciepts", "6M Excel Used", "7M Excel Reciepts", "7M Excel Used", "8M Excel Reciepts", "8M Excel Used", "9M Excel Reciepts", "9M Excel Used", "10M Excel Reciepts", "10M Excel Used", "12M Excel Reciepts", "12M Excel Used", "15M Excel Reciepts", "15M Excel Used", "Slurry Reciept", "Slurry Used", "Slurry Big Reciept", "Slurry Big Used", "E.D Reciept", "E.D Used", "D.Det Reciepts", "D.Det Used", "DF 5gms Reciepts", "DF 5gms Used", "DF 10gms Reciepts", "DF 10gms Used", "No of Holes 2.5", "No of Holes 5", "No of Holes 6", "No of Holes 8", "Wagon Drill 4.5''", "Amount Paid to Balan"]
        data2=[date1,blasting_done_on,_3m_excel_,_3m_excel_used_,_4m_reciepts_,_4m_excel_used_,_5m_excel_reciepts_,_5m_excel_used_,_6m_excel_reciepts_,_6m_excel_used_,_7m_excel_reciepts_,_7m_excel_used_,_8m_excel_reciepts_,_8m_excel_used_,_9m_excel_reciepts_,_9m_excel_used_,_10m_excel_reciepts_,_10m_excel_used_,_12m_excel_reciepts_,_12m_excel_used_,_15m_excel_reciepts_,_15m_excel_used_,_slurry_reciept_,_slurry_used_,_slurry_big_reciept_,_slurry_big_used_,_ed_reciept_,_ed_used_,_d_det_reciepts_,_d_det_used_,df_5gms_reciepts_,df_5gms_used_,_df_10gms_reciepts_,_df_10gms_used_,_no_of_holes_2_5_,_no_of_holes_5_,_no_of_holes_6_,_no_of_holes_8_,_wagon_drill_4_5_,amount_paid_]
        global filepath
        if any(field == '' or field == ' ' for field in data2):
            messagebox.showerror("Error", "Please Fill All the Fields.")
        else:
            if not os.path.exists(filepath):
                workbook = openpyxl.Workbook()
                workbook.create_sheet("Explosives")
                sheet=workbook.active=workbook["Explosives"]
                sheet.append(heading)
                sheet.append(data2)
                workbook.save(filepath)
                clear1()
                messagebox.showinfo("Success","File Created and Data Submitted Successfully")
            else:
                workbook = openpyxl.load_workbook(filepath)
                #Check if sheet exists or not
                if "Explosives" not in workbook.sheetnames:
                    workbook.create_sheet("Explosives")
                    sheet=workbook.active=workbook["Explosives"]
                    sheet.append(heading)            
                    sheet.append(data2)
                    workbook.save(filepath)
                    clear1()
                    messagebox.showinfo("Success","Sheet Created and Submitted Succesfully.")
                else:
                    sheet=workbook["Explosives"]
                    sheet.append(data2)
                    workbook.save(filepath)
                    clear1()
                    messagebox.showinfo("Success","Data Submitted Successfully")

    def edit1():
        global window1
        global date_edit_entry
        global filepath
        window1 = tkinter.Tk()
        window1.title("Edit Data")
        window1.configure(background="#CCCCFF")
        frame1 = ttk.Frame(window1, padding=20)
        frame1.pack()
        dlabel = ttk.Label(frame1, text="Enter Date to Edit:", background="#CCCCFF")
        dlabel.grid(row=0, column=0)
        date_edit_entry = ttk.Entry(frame1)
        date_edit_entry.grid(row=0, column=3)
        blasting_done_on_label_ = ttk.Label(frame1, text="Blasting Done On", background="#CCCCFF")
        blasting_done_on_label_.grid(row=1, column=0)
        blasting_done_on_edit_entry_ = ttk.Combobox(frame1, values=["Soft Rock", "Hard Rock"])
        blasting_done_on_edit_entry_.grid(row=1, column=3)
        def edit_data():
            date_edit = date_edit_entry.get()

            workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active = workbook["Explosives"]
            found = False
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True):
                if row[0] == date_edit and row[1] == blasting_done_on_edit_entry_.get():
                    fill_data(row[0:])
                    found = True
                    break
            if found:
                messagebox.showinfo("Success", "Data Found and Filled in the Form.")
            else:
                messagebox.showinfo("Error", "Data Not Found.")
            window1.destroy()

        def fill_data(data):
            entries = [
                date1_entry, blasting_done_on_entry_, _3m_excel_entry_, _3m_excel_used_entry_, _4m_reciepts_entry_, _4m_excel_used_entry_,
                _5m_excel_reciepts_entry_, _5m_excel_used_entry_, _6m_excel_reciepts_entry_, _6m_excel_used_entry_, _7m_excel_reciepts_entry_,
                _7m_excel_used_entry_, _8m_excel_reciepts_entry_, _8m_excel_used_entry_, _9m_excel_reciepts_entry_, _9m_excel_used_entry_,
                _10m_excel_reciepts_entry_, _10m_excel_used_entry_, _12m_excel_reciepts_entry_, _12m_excel_used_entry_, _15m_excel_reciepts_entry_,
                _15m_excel_used_entry_, _slurry_reciept_entry_, _slurry_used_entry_, _slurry_big_reciept_entry_, _slurry_big_used_entry_,
                _ed_reciept_entry_, _ed_used_entry_, _d_det_reciepts_entry_, _d_det_used_entry_, df_5gms_reciepts_entry_, df_5gms_used_entry_,
                _df_10gms_reciepts_entry_, _df_10gms_used_entry_, _no_of_holes_2_5_entry_, _no_of_holes_5_entry_, _no_of_holes_6_entry_,
                _no_of_holes_8_entry_, _wagon_drill_4_5_entry_, amount_paid_entry_
                ]

            for entry, value in zip(entries, data):
                entry.insert(0, value)

        button = ttk.Button(frame1, text="Edit", command=edit_data)
        button.grid(row=2, column=1)
        window1.iconbitmap(basepath+"\\906310.ico")
        window1.mainloop()
    def modify1():
        global window1
        global filepath
        workbook = openpyxl.load_workbook(filepath)
        sheet=workbook["Explosives"]
        date=date1_entry.get()
        blasting_done_on=blasting_done_on_entry_.get()
        _3m_excel_=float(_3m_excel_entry_.get())
        _3m_excel_used_=float(_3m_excel_used_entry_.get())
        _4m_reciepts_=float(_4m_reciepts_entry_.get())
        _4m_excel_used_=float(_4m_excel_used_entry_.get())
        _5m_excel_reciepts_=float(_5m_excel_reciepts_entry_.get())
        _5m_excel_used_=float(_5m_excel_used_entry_.get())
        _6m_excel_reciepts_=float(_6m_excel_reciepts_entry_.get())
        _6m_excel_used_=float(_6m_excel_used_entry_.get())
        _7m_excel_reciepts_=float(_7m_excel_reciepts_entry_.get())
        _7m_excel_used_=float(_7m_excel_used_entry_.get())
        _8m_excel_reciepts_=float(_8m_excel_reciepts_entry_.get())
        _8m_excel_used_=float(_8m_excel_used_entry_.get())
        _9m_excel_reciepts_=float(_9m_excel_reciepts_entry_.get())
        _9m_excel_used_=float(_9m_excel_used_entry_.get())
        _10m_excel_reciepts_=float(_10m_excel_reciepts_entry_.get())
        _10m_excel_used_=float(_10m_excel_used_entry_.get())
        _12m_excel_reciepts_=float(_12m_excel_reciepts_entry_.get())
        _12m_excel_used_=float(_12m_excel_used_entry_.get())
        _15m_excel_reciepts_=float(_15m_excel_reciepts_entry_.get())
        _15m_excel_used_=float(_15m_excel_used_entry_.get())
        _slurry_reciept_=float(_slurry_reciept_entry_.get())
        _slurry_used_=float(_slurry_used_entry_.get())
        _slurry_big_reciept_=float(_slurry_big_reciept_entry_.get())
        _slurry_big_used_=float(_slurry_big_used_entry_.get())
        _ed_reciept_=float(_ed_reciept_entry_.get())
        _ed_used_=float(_ed_used_entry_.get())
        _d_det_reciepts_=float(_d_det_reciepts_entry_.get())
        _d_det_used_=float(_d_det_used_entry_.get())
        df_5gms_reciepts_=float(df_5gms_reciepts_entry_.get())
        df_5gms_used_=float(df_5gms_used_entry_.get())
        _df_10gms_reciepts_=float(_df_10gms_reciepts_entry_.get())
        _df_10gms_used_=float(_df_10gms_used_entry_.get())
        _no_of_holes_2_5_=float(_no_of_holes_2_5_entry_.get())
        _no_of_holes_5_=float(_no_of_holes_5_entry_.get())
        _no_of_holes_6_=float(_no_of_holes_6_entry_.get())
        _no_of_holes_8_=float(_no_of_holes_8_entry_.get())
        _wagon_drill_4_5_=float(_wagon_drill_4_5_entry_.get())
        amount_paid_=float(amount_paid_entry_.get())
        if any(field == '' or field == ' ' for field in [date,blasting_done_on,_3m_excel_,_3m_excel_used_,_4m_reciepts_,_4m_excel_used_,_5m_excel_reciepts_,_5m_excel_used_,_6m_excel_reciepts_,_6m_excel_used_,_7m_excel_reciepts_,_7m_excel_used_,_8m_excel_reciepts_,_8m_excel_used_,_9m_excel_reciepts_,_9m_excel_used_,_10m_excel_reciepts_,_10m_excel_used_,_12m_excel_reciepts_,_12m_excel_used_,_15m_excel_reciepts_,_15m_excel_used_,_slurry_reciept_,_slurry_used_,_slurry_big_reciept_,_slurry_big_used_,_ed_reciept_,_ed_used_,_d_det_reciepts_,_d_det_used_,df_5gms_reciepts_,df_5gms_used_,_df_10gms_reciepts_,_df_10gms_used_,_no_of_holes_2_5_,_no_of_holes_5_,_no_of_holes_6_,_no_of_holes_8_,_wagon_drill_4_5_,amount_paid_]):
            messagebox.showerror("Error", "Please Fill All the Fields.")
        else:
            for i in range(1,sheet.max_row+1):
                if sheet.cell(row=i,column=1).value==date:
                    sheet.cell(row=i,column=2).value=blasting_done_on
                    sheet.cell(row=i,column=3).value=_3m_excel_
                    sheet.cell(row=i,column=4).value=_3m_excel_used_
                    sheet.cell(row=i,column=5).value=_4m_reciepts_
                    sheet.cell(row=i,column=6).value=_4m_excel_used_
                    sheet.cell(row=i,column=7).value=_5m_excel_reciepts_
                    sheet.cell(row=i,column=8).value=_5m_excel_used_
                    sheet.cell(row=i,column=9).value=_6m_excel_reciepts_
                    sheet.cell(row=i,column=10).value=_6m_excel_used_
                    sheet.cell(row=i,column=11).value=_7m_excel_reciepts_
                    sheet.cell(row=i,column=12).value=_7m_excel_used_
                    sheet.cell(row=i,column=13).value=_8m_excel_reciepts_
                    sheet.cell(row=i,column=14).value=_8m_excel_used_
                    sheet.cell(row=i,column=15).value=_9m_excel_reciepts_
                    sheet.cell(row=i,column=16).value=_9m_excel_used_
                    sheet.cell(row=i,column=17).value=_10m_excel_reciepts_
                    sheet.cell(row=i,column=18).value=_10m_excel_used_
                    sheet.cell(row=i,column=19).value=_12m_excel_reciepts_
                    sheet.cell(row=i,column=20).value=_12m_excel_used_
                    sheet.cell(row=i,column=21).value=_15m_excel_reciepts_
                    sheet.cell(row=i,column=22).value=_15m_excel_used_  
                    sheet.cell(row=i,column=23).value=_slurry_reciept_
                    sheet.cell(row=i,column=24).value=_slurry_used_
                    sheet.cell(row=i,column=25).value=_slurry_big_reciept_
                    sheet.cell(row=i,column=26).value=_slurry_big_used_
                    sheet.cell(row=i,column=27).value=_ed_reciept_
                    sheet.cell(row=i,column=28).value=_ed_used_
                    sheet.cell(row=i,column=29).value=_d_det_reciepts_
                    sheet.cell(row=i,column=30).value=_d_det_used_
                    sheet.cell(row=i,column=31).value=df_5gms_reciepts_
                    sheet.cell(row=i,column=32).value=df_5gms_used_
                    sheet.cell(row=i,column=33).value=_df_10gms_reciepts_
                    sheet.cell(row=i,column=34).value=_df_10gms_used_
                    sheet.cell(row=i,column=35).value=_no_of_holes_2_5_
                    sheet.cell(row=i,column=36).value=_no_of_holes_5_
                    sheet.cell(row=i,column=37).value=_no_of_holes_6_
                    sheet.cell(row=i,column=38).value=_no_of_holes_8_
                    sheet.cell(row=i,column=39).value=_wagon_drill_4_5_
                    sheet.cell(row=i,column=40).value=amount_paid_
                    workbook.save(filepath)
                    clear1()
                    break
            messagebox.showinfo("Success","Data Modified Successfully")
    def delete1():
        global filepath
        window2 = tkinter.Tk()
        window2.title("Delete Data")
        window2.configure(bg="#CCCCFF")
        frame1 = ttk.Frame(window2,padding=20)
        frame1.pack()
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook["Explosives"]
        dlabel = ttk.Label(frame1, text="Enter Date to Delete:", background="#CCCCFF")
        dlabel.grid(row=0, column=0)
        date_delete_entry = ttk.Entry(frame1)
        date_delete_entry.grid(row=0, column=1)
        blasting_done_delete_entry_=ttk.Combobox(frame1,values=["Soft Rock","Hard Rock"])
        blasting_done_delete_entry_.grid(row=0,column=2)
        def delete_data1():
            date_delete = date_delete_entry.get()
            found = False
            for i in range(1, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == date_delete and sheet.cell(row=i,column=2).value==blasting_done_delete_entry_.get():
                    sheet.delete_rows(i)
                    workbook.save(filepath)
                    found = True
                    break
            window2.destroy()
            if found:
                messagebox.showinfo("Success", "Data Deleted Successfully")
            else:
                messagebox.showinfo("Error", "Data Not Found")
        button_delete1 = ttk.Button(frame1, text="Delete", command=delete_data1)
        button_delete1.grid(row=3, column=0)
        window2.iconbitmap(basepath+"\\906310.ico")
        window2.mainloop()
    # Button Frame
    button1_frame = tkinter.Frame(second_frame)
    button1_frame.grid(row= 3, column=0, padx=40, pady=10)
    # Button Save
    button1_save = tkinter.Button(button1_frame, text="ㅤㅤㅤSaveㅤㅤㅤ", command= save)
    button1_save.configure(background="green")
    button1_save.grid(row=0, column=0, sticky="news")
    # Button Clear
    button1_clear = tkinter.Button(button1_frame, text="ㅤㅤㅤClearㅤㅤㅤ", command= clear1)
    button1_clear.configure(background="yellow")
    button1_clear.grid(row=0, column=1, sticky="news")
    # Button Edit
    button1_edit = tkinter.Button(button1_frame, text="ㅤㅤㅤEditㅤㅤㅤ", command= edit1)
    button1_edit.configure(background="cyan")
    button1_edit.grid(row=0, column=2, sticky="news")
    #Button Delete
    button1_delete = tkinter.Button(button1_frame, text="ㅤㅤㅤDeleteㅤㅤㅤ", command= delete1)
    button1_delete.configure(background="red")
    button1_delete.grid(row=0, column=4, sticky="news")
    #Button Modify
    button1_modify=tkinter.Button(button1_frame,text="ㅤㅤㅤModifyㅤㅤㅤ",command=modify1)
    button1_modify.configure(background="orange")
    button1_modify.grid(row=0,column=3,sticky="news")
    #-----------------------------------------------------------End of Explosives---------------------------------------------------------#
    # #-----------------------------------------------------------Code for Displayng Stock--------------------------------------------------#
    # data2_info = ttk.LabelFrame(second_frame, text="Stock", padding=20)
    # data2_info.grid(row=4, column=0, padx=40, pady=10)
    # #Display Stock
    # display_label=ttk.Label(data2_info,text="Display Stock : ")
    # display_label.grid(row=0,column=0)
    # display_entry=ttk.Combobox(data2_info,values=["Total","Date Wise"])
    # display_entry.grid(row=0,column=1)      
    # explosive_type_label = ttk.Label(data2_info, text="Name of The Stock to show : ")
    # explosive_type_label.grid(row=1, column=0)
    # explosive_type_entry = ttk.Combobox(data2_info, values=["4M", "5M", "Slurry", "Slurry Big", "ED", "3M", "D.Det", "DF 10gms", "DF 5gms"])
    # explosive_type_entry.grid(row=1, column=1)
    # if display_entry.get()=="Date Wise":
    #     window3=tkinter.Tk()
    #     window3.title("Date Wise")
    #     window3.configure(bg="#CCCCFF")
    #     frame1=ttk.Frame(window3,padding=20)
    #     datestock_label=ttk.Label(frame1,text="Date : ")
    #     datestock_label.grid(row=0,column=0)
    #     datestock_entry=ttk.Entry(frame1)
    #     datestock_entry.grid(row=0,column=1)    
    #     datestockbutton=ttk.Button(frame1,text="Show",command=get_stocks)
    #     datestockbutton.grid(row=1,column=0)
    #     window3.iconbitmap(basepath+"\\906310.ico")
    #     window3.mainloop()
    # #Stock Mechanism
    # def get_stocks():
    #     filepath="D://Excel//mpdrentry.xlsx"
    #     global display_entry
    #     global date2_entry
    #     global explosive_type_entry
    #     workbook=openpyxl.load_workbook(filepath)
    #     sheet=workbook["Explosives"]
    #     _4mtotal=0;_4mused=0
    #     _5mtotal=0;_5mused=0
    #     slurrytotal=0;slurryused=0
    #     slurrybigtotal=0;slurrybigused=0
    #     edtotal=0;edused=0
    #     _3mtotal=0;_3mused=0
    #     d_dettotal=0;d_detused=0
    #     df_10gmstotal=0;df_10gmsused=0
    #     df_5gmstotal=0;df_5gmsused=0
    #     if explosive_type_entry.get()=="4M":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     _4mstock += float(row[2])  # Opening Balance
    #                     _4mused += float(row[3])  # Used
    #             else:
    #                 _4mstock += float(row[2])  # Opening Balance
    #                 _4mused += float(row[3])  # Used
    #         _4mstock=_4mtotal-_4mused
    #         return _4mstock
    #     elif explosive_type_entry.get()=="5M":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     _5mtotal += float(row[4])
    #                     _5mused += float(row[5])
    #             else:
    #                 _5mtotal += float(row[4])
    #                 _5mused += float(row[5])
    #         _5mstock=_5mtotal-_5mused
    #         return _5mstock
    #     elif explosive_type_entry.get()=="Slurry":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     slurrytotal += float(row[6])
    #                     slurryused += float(row[7])
    #             else:
    #                 slurrytotal += float(row[6])
    #                 slurryused += float(row[7])
    #         slurrystock=slurrytotal-slurryused
    #         return slurrystock
    #     elif explosive_type_entry.get()=="Slurry Big":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     slurrybigtotal += float(row[8])
    #                     slurrybigused += float(row[9])
    #             else:
    #                 slurrybigtotal += float(row[8])
    #                 slurrybigused += float(row[9])
    #         slurrybigstock=slurrybigtotal-slurrybigused
    #         return slurrybigstock
    #     elif explosive_type_entry.get()=="ED":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     edtotal += float(row[10])
    #                     edused += float(row[11])
    #             else:
    #                 edtotal += float(row[10])
    #                 edused += float(row[11])
    #         edstock=edtotal-edused
    #         return edstock
    #     elif explosive_type_entry.get()=="3M":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     _3mtotal += float(row[12])
    #                     _3mused += float(row[13])
    #             else:
    #                 _3mtotal += float(row[12])
    #                 _3mused += float(row[13])
    #         _3mstock=_3mtotal-_3mused
    #         return _3mstock
    #     elif explosive_type_entry.get()=="D.Det":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     d_dettotal += float(row[14])
    #                     d_detused += float(row[15])
    #             else:
    #                 d_dettotal += float(row[14])
    #                 d_detused += float(row[15])
    #         d_detstock=d_dettotal-d_detused
    #         return d_detstock
    #     elif explosive_type_entry.get()=="DF 10gms":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     df_10gmstotal += float(row[16])
    #                     df_10gmsused += float(row[17])
    #             else:
    #                 df_10gmstotal += float(row[16])
    #                 df_10gmsused += float(row[17])
    #         df_10gmsstock=df_10gmstotal-df_10gmsused
    #         return df_10gmsstock
    #     elif explosive_type_entry.get()=="DF 5gms":
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             if display_entry == "Date Wise":
    #                 if row[0] == datestock_entry.get():
    #                     df_5gmstotal += float(row[24])
    #                     df_5gmsused += float(row[25])
    #             else:
    #                 df_5gmstotal += float(row[24])
    #                 df_5gmsused += float(row[25])
    #         df_5gmsstock=df_5gmstotal-df_5gmsused
    #         return df_5gmsstock
    #     else:
    #         messagebox.showerror("Error","Please Select a Type of Stock.")
    # value=0
    # stock_label = ttk.Label(data2_info, text="Select the option and click on Get Info.")
    # stock_label.grid(row=4, column=0)
    # def update_stocks():
    #     global value
    #     value=get_stocks()
    #     stock_label.configure(text=f"Stock : {value}")
    # button_get_info=ttk.Button(data2_info,text="Get Info",command=update_stocks)
    # button_get_info.grid(row=2,column=0)
    # # def get_stock_info():
    # #     filepath = "D://Excel//mpdrentry.xlsx"
    # #     # Retrieve the selected options from the comboboxes
    # #     display_option = display_entry.get()
    # #     explosive_type = explosive_type_entry.get()

    # #     # Open the workbook and select the appropriate sheet
    # #     workbook = openpyxl.load_workbook(filepath)
    # #     sheet = workbook["Explosives"]

    # #     # Perform calculations based on the selected options
    # #     total_stock = 0
    # #     total_used = 0
    # #     for row in sheet.iter_rows(min_row=2, values_only=True):
    # #         if display_option == "Date Wise":
    # #             if row[0] == date_entry.get():
    # #                 total_stock += float(row[2])  # Opening Balance
    # #                 total_used += float(row[3])  # Used
    # #         else:
    # #             total_stock += float(row[2])  # Opening Balance
    # #             total_used += float(row[3])  # Used

    # #     # Calculate the remaining stock
    # #     remaining_stock = total_stock - total_used

    # #     # Update the stock label
    # #     stock_label.config(text=f"Remaining Stock: {remaining_stock}")
    # # # Create the labels
    # # display_label = ttk.Label(data2_info, text="Display Stock:")
    # # display_label.grid(row=0, column=0, padx=10, pady=5)

    # # explosive_type_label = ttk.Label(data2_info, text="Explosive Type:")
    # # explosive_type_label.grid(row=1, column=0, padx=10, pady=5)

    # # # Create the comboboxes
    # # display_entry = ttk.Combobox(data2_info, values=["Total", "Date Wise"])
    # # display_entry.grid(row=0, column=1, padx=10, pady=5)

    # # explosive_type_entry = ttk.Combobox(data2_info, values=["4M", "5M", "Slurry", "Slurry Big", "ED", "3M", "D.Det", "DF 10gms", "DF 5gms"])
    # # explosive_type_entry.grid(row=1, column=1, padx=10, pady=5)

    # # # Create the stock label
    # # stock_label = ttk.Label(data2_info, text="Remaining Stock: 0")
    # # stock_label.grid(row=2, column=0, columnspan=2, padx=10, pady=5)

    # # # Create the button to retrieve stock information
    # # get_info_button = ttk.Button(data2_info, text="Get Info", command=get_stock_info)
    # # get_info_button.grid(row=3, column=0, columnspan=2, padx=10, pady=5)

    # #-----------------------------------------------------------End of Code for Displaying Stock------------------------------------------#
    # #Window Logo
    window.iconbitmap(basepath+"\\906310.ico")
    #MainLoop
    window.mainloop()
    #------------------------------------------------------------End of Program-----------------------------------------------------------#
